#!/usr/bin/env python3
"""
ID Card Exporter Script

This script connects to Supabase, fetches ID card data, downloads high-quality images,
and creates an Excel file with all details including serial numbers as admission numbers.

Usage:
    python export_id_cards.py [--output OUTPUT_DIR] [--start-serial START_SERIAL]

Options:
    --output OUTPUT_DIR      Directory to save the Excel file and images (default: ./output)
    --start-serial START     Starting serial number for admission numbers (default: 115601)
    --class-id CLASS_ID      Filter by class ID (optional)
    --search SEARCH          Search term for filtering (optional)
"""

import os
import sys
import argparse
import json
import requests
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import supabase
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Constants
SCHEMA = "school"
ID_CARD_TABLE = "IDCard"
CLASS_TABLE = "Class"
STORAGE_BUCKET = "File"

# Get Supabase credentials from environment variables
SUPABASE_URL = os.getenv("VITE_SUPABASE_URL")
SUPABASE_KEY = os.getenv("VITE_SUPABASE_ANON_KEY")
SUPABASE_SERVICE_KEY = os.getenv("VITE_SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not (SUPABASE_KEY or SUPABASE_SERVICE_KEY):
    print("Error: Supabase credentials not found in environment variables.")
    print("Please set VITE_SUPABASE_URL and either VITE_SUPABASE_ANON_KEY or VITE_SUPABASE_SERVICE_ROLE_KEY.")
    sys.exit(1)

# Use service key if available, otherwise use anon key
SUPABASE_API_KEY = SUPABASE_SERVICE_KEY if SUPABASE_SERVICE_KEY else SUPABASE_KEY

# Create Supabase client
client = supabase.create_client(SUPABASE_URL, SUPABASE_API_KEY)

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Export ID cards with high-quality images")
    parser.add_argument("--output", default="./output", help="Output directory for Excel and images")
    parser.add_argument("--start-serial", type=int, default=115601, help="Starting serial number for admission numbers")
    parser.add_argument("--class-id", help="Filter by class ID")
    parser.add_argument("--search", help="Search term for filtering")
    return parser.parse_args()

def ensure_output_directory(output_dir):
    """Ensure the output directory exists."""
    path = Path(output_dir)
    images_path = path / "images"
    
    # Create directories if they don't exist
    path.mkdir(exist_ok=True)
    images_path.mkdir(exist_ok=True)
    
    return path, images_path

def get_class_map():
    """Get a mapping of class IDs to class names and sections."""
    response = client.table(f"{SCHEMA}.{CLASS_TABLE}").select("id,name,section").execute()
    
    if hasattr(response, 'error') and response.error:
        print(f"Error fetching classes: {response.error}")
        return {}
    
    class_map = {}
    for cls in response.data:
        class_map[cls['id']] = {
            'name': cls.get('name', 'Unknown'),
            'section': cls.get('section', '')
        }
    
    return class_map

def fetch_id_cards(class_id=None, search=None):
    """Fetch ID cards from Supabase with optional filtering."""
    query = client.table(f"{SCHEMA}.{ID_CARD_TABLE}").select("""
        id,
        student_name,
        class_id,
        date_of_birth,
        student_photo_url,
        father_name,
        mother_name,
        father_photo_url,
        mother_photo_url,
        father_mobile,
        mother_mobile,
        address,
        created_at,
        download_count
    """)
    
    # Apply filters if provided
    if class_id:
        query = query.eq('class_id', class_id)
    
    if search:
        query = query.or_(
            f"student_name.ilike.%{search}%,father_name.ilike.%{search}%,mother_name.ilike.%{search}%,address.ilike.%{search}%"
        )
    
    response = query.execute()
    
    if hasattr(response, 'error') and response.error:
        print(f"Error fetching ID cards: {response.error}")
        return []
    
    return response.data

def download_image(url, output_path):
    """Download an image from a URL and save it to the specified path."""
    if not url or not url.strip():
        return None
    
    try:
        response = requests.get(url, stream=True)
        if response.status_code == 200:
            with open(output_path, 'wb') as f:
                for chunk in response.iter_content(1024):
                    f.write(chunk)
            return output_path
        else:
            print(f"Failed to download image: {url}, Status code: {response.status_code}")
            return None
    except Exception as e:
        print(f"Error downloading image {url}: {e}")
        return None

def sanitize_filename(name):
    """Sanitize a string to be used as a filename."""
    if not name:
        return "unknown"
    
    # Replace invalid characters with underscores
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Limit length and remove leading/trailing spaces
    return name.strip()[:50]

def create_excel_file(id_cards, class_map, output_dir, images_dir, start_serial):
    """Create an Excel file with ID card data and high-quality images."""
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "ID Cards"
    
    # Define headers
    headers = [
        "Serial No.",
        "Admission No.",
        "Student Photo",
        "Student Name",
        "Class",
        "Date of Birth",
        "Father Photo",
        "Father Name",
        "Father Mobile",
        "Mother Photo",
        "Mother Name",
        "Mother Mobile",
        "Address",
        "Created Date"
    ]
    
    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Set column widths
    column_widths = {
        1: 10,  # Serial No.
        2: 15,  # Admission No.
        3: 15,  # Student Photo
        4: 25,  # Student Name
        5: 15,  # Class
        6: 15,  # Date of Birth
        7: 15,  # Father Photo
        8: 25,  # Father Name
        9: 15,  # Father Mobile
        10: 15,  # Mother Photo
        11: 25,  # Mother Name
        12: 15,  # Mother Mobile
        13: 40,  # Address
        14: 15,  # Created Date
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Process each ID card
    image_paths = []
    for idx, card in enumerate(id_cards, 1):
        row_num = idx + 1
        serial_num = start_serial + idx - 1
        admission_num = f"ADM{serial_num}"
        
        # Get class details
        class_info = class_map.get(card.get('class_id'), {'name': 'Unknown', 'section': ''})
        class_name = f"{class_info['name']} {class_info['section']}".strip()
        
        # Format date of birth
        dob = card.get('date_of_birth')
        if dob:
            try:
                dob = datetime.fromisoformat(dob.replace('Z', '+00:00')).strftime('%d-%m-%Y')
            except:
                dob = dob
        
        # Format created date
        created_at = card.get('created_at')
        if created_at:
            try:
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00')).strftime('%d-%m-%Y')
            except:
                created_at = created_at
        
        # Download and save images
        student_name = sanitize_filename(card.get('student_name', ''))
        father_name = sanitize_filename(card.get('father_name', ''))
        mother_name = sanitize_filename(card.get('mother_name', ''))
        
        # Create image filenames
        student_img_filename = f"{serial_num}_{student_name}_student.jpg"
        father_img_filename = f"{serial_num}_{student_name}_{father_name}_father.jpg"
        mother_img_filename = f"{serial_num}_{student_name}_{mother_name}_mother.jpg"
        
        # Download images
        student_img_path = None
        father_img_path = None
        mother_img_path = None
        
        if card.get('student_photo_url'):
            student_img_path = download_image(
                card['student_photo_url'], 
                os.path.join(images_dir, student_img_filename)
            )
            if student_img_path:
                image_paths.append(student_img_path)
        
        if card.get('father_photo_url'):
            father_img_path = download_image(
                card['father_photo_url'], 
                os.path.join(images_dir, father_img_filename)
            )
            if father_img_path:
                image_paths.append(father_img_path)
        
        if card.get('mother_photo_url'):
            mother_img_path = download_image(
                card['mother_photo_url'], 
                os.path.join(images_dir, mother_img_filename)
            )
            if mother_img_path:
                image_paths.append(mother_img_path)
        
        # Add data to worksheet
        ws.cell(row=row_num, column=1).value = idx  # Serial No.
        ws.cell(row=row_num, column=2).value = admission_num  # Admission No.
        ws.cell(row=row_num, column=3).value = student_img_filename if student_img_path else "N/A"  # Student Photo
        ws.cell(row=row_num, column=4).value = card.get('student_name', '')  # Student Name
        ws.cell(row=row_num, column=5).value = class_name  # Class
        ws.cell(row=row_num, column=6).value = dob  # Date of Birth
        ws.cell(row=row_num, column=7).value = father_img_filename if father_img_path else "N/A"  # Father Photo
        ws.cell(row=row_num, column=8).value = card.get('father_name', '')  # Father Name
        ws.cell(row=row_num, column=9).value = card.get('father_mobile', '')  # Father Mobile
        ws.cell(row=row_num, column=10).value = mother_img_filename if mother_img_path else "N/A"  # Mother Photo
        ws.cell(row=row_num, column=11).value = card.get('mother_name', '')  # Mother Name
        ws.cell(row=row_num, column=12).value = card.get('mother_mobile', '')  # Mother Mobile
        ws.cell(row=row_num, column=13).value = card.get('address', '')  # Address
        ws.cell(row=row_num, column=14).value = created_at  # Created Date
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(id_cards) + 1):
        for cell in row:
            cell.border = thin_border
    
    # Add auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(id_cards) + 1}"
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"ID_Cards_Export_{timestamp}.xlsx")
    wb.save(excel_path)
    
    return excel_path, image_paths

def main():
    """Main function to run the script."""
    args = parse_arguments()
    
    # Ensure output directory exists
    output_dir, images_dir = ensure_output_directory(args.output)
    
    print(f"Connecting to Supabase at {SUPABASE_URL}...")
    
    # Get class mapping
    print("Fetching class information...")
    class_map = get_class_map()
    
    # Fetch ID cards
    print("Fetching ID card data...")
    id_cards = fetch_id_cards(args.class_id, args.search)
    
    if not id_cards:
        print("No ID cards found with the specified criteria.")
        return
    
    print(f"Found {len(id_cards)} ID cards. Processing...")
    
    # Create Excel file with images
    print("Creating Excel file and downloading images...")
    excel_path, image_paths = create_excel_file(
        id_cards, 
        class_map, 
        output_dir, 
        images_dir, 
        args.start_serial
    )
    
    print(f"\nExport completed successfully!")
    print(f"Excel file saved to: {excel_path}")
    print(f"Downloaded {len(image_paths)} images to: {images_dir}")
    print(f"Total ID cards processed: {len(id_cards)}")

if __name__ == "__main__":
    main()
